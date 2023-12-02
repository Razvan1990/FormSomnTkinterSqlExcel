import sqlite3
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Side, PatternFill, Alignment, Font
from tkinter import messagebox
from openpyxl.utils import get_column_letter
import constants_pacienti


class ExcelWriter:

    def get_max_value(self, list_text):
        max = list_text[0]
        for i in range(1, len(list_text)):
            if list_text[i] > max:
                max = list_text[i]
        return max

    def write_to_excel(self, table_name):
        excel_location = os.path.join(os.getcwd(), constants_pacienti.EXCEL_FOLDER, constants_pacienti.NAME_EXCEL)
        # os.chdir(excel_location)
        '''SQL SELECT'''
        database = os.path.join(os.getcwd(), constants_pacienti.DATABASE_FOLDER, constants_pacienti.NAME_DATABASE)
        connection = sqlite3.connect(database)
        my_cursor = connection.cursor()
        # my_cursor.execute("""SELECT * FROM """+table_name)
        # list_patients = my_cursor.fetchall()
        sql = "SELECT * FROM " + table_name
        # start with excel
        # df_excel = pd.DataFrame(list_patients, columns=constants_pacienti.SQL_HEADERS)
        # df_excel.to_excel(excel_location, sheet_name="REGISTRU")
        # wb = openpyxl.load_workbook(filename=excel_location)
        # work_sheet = wb["REGISTRU"]
        try:
            df = pd.read_sql(sql, connection)
            df.to_excel(excel_location)
            wb = openpyxl.load_workbook(filename=excel_location)
            work_sheet = wb["Sheet1"]
            work_sheet.title = "REGISTRU"
            color = openpyxl.styles.colors.Color(rgb="00F5FFDE")
            custom_fill = PatternFill(fill_type="solid", fgColor=color)
            # all columns with excel data
            list_index_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
                                  "R",
                                  "S", "T", "U", "V"]
            # MAKE THE FIRST COLUMN AS AN ID
            work_sheet["A1"].value = "ID"
            work_sheet["A1"].alignment = Alignment(horizontal="center")
            work_sheet["A1"].font = Font(bold=True)
            for column in list_index_columns:
                # make a custome style
                work_sheet["{}1".format(column)].fill = custom_fill
            '''wrap text'''
            # first let's see what is the max value of each row
            list_values_column = []
            for index_column in range(1, len(list_index_columns)):
                list_length_column = list()
                for i in range(1, work_sheet.max_row + 1):
                    # in case it is not completed the cell
                    if work_sheet["{}{}".format(list_index_columns[index_column], i)].value == None:
                        continue
                    else:
                        column_row_length = len(work_sheet["{}{}".format(list_index_columns[index_column], i)].value)
                        list_length_column.append(column_row_length)
                max_value_column = self.get_max_value(list_length_column)
                list_values_column.append(max_value_column)
            # now traverse every column and make than column the width of the max value
            index = 0
            for index_column in range(1, len(list_index_columns)):
                work_sheet.column_dimensions[list_index_columns[index_column]].width = list_values_column[
                                                                                           index] + 10  # make them a bit bigger
                index += 1
            # apply filters => stack overflow code
            full_range = "A1:" + get_column_letter(work_sheet.max_column) \
                         + str(work_sheet.max_row)
            work_sheet.auto_filter.ref = full_range
            wb.save(excel_location)
            message_excel = "Baza de date cu pacienti este transferata si disponibila pe {}".format(excel_location)
            messagebox.showinfo("EXCEL CREATED", message=message_excel)
            os.system(excel_location)
        except:
            messagebox.showerror("INCHIDETI FISIER EXCEL",
                                 "Fisierul Registru Pacienti este deschis! Va rog inchideti-l")
            raise Exception("EXCEL file is already opened")
