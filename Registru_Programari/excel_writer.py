import sqlite3
import pandas as pd
import openpyxl
import os
import constants_programari
from openpyxl.styles import Side, PatternFill, Alignment, Font
from tkinter import messagebox
from openpyxl.utils import get_column_letter
from datetime import datetime
from checker_fields import CheckFields


class ExcelWriter:

    def __init__(self):
        self.checker = CheckFields()
        self.excel_location = os.path.join(os.getcwd(), constants_programari.EXCEL_FOLDER,
                                           constants_programari.NAME_EXCEL)

    def get_max_value(self, list_text):
        max = list_text[0]
        for i in range(1, len(list_text)):
            if list_text[i] > max:
                max = list_text[i]
        return max

    def write_to_excel(self):
        '''
        here we will use a sql statement first to get all the tables
        :return: An Excel with all tables
        '''
        # excel_location = os.path.join(os.getcwd(), constants_programari.EXCEL_FOLDER, constants_programari.NAME_EXCEL)
        try:
            writer = pd.ExcelWriter(self.excel_location, engine="xlsxwriter")
            database = os.path.join(os.getcwd(), constants_programari.DATABASE_FOLDER, constants_programari.NAME_DATABASE)
            connection = sqlite3.connect(database)
            my_cursor = connection.cursor()
            # sql command found
            sql_retrieve_table_command = """SELECT name FROM sqlite_schema WHERE type ='table' """
            my_cursor.execute(sql_retrieve_table_command)
            list_tables_sql = my_cursor.fetchall()
            # get list of tables from returned list of tuples
            list_tables_final = list()  # this list is needed to retrieve from sql
            for tuple_name in list_tables_sql:
                list_tables_final.append(tuple_name[0])
            # transform list with date format
            list_tables_final_dates = list()
            for name_date in list_tables_final:
                date = self.checker.reconvert_date(name_date)
                list_tables_final_dates.append(date)
            # sort the list in cronological order
            list_tables_final_dates.sort(key=lambda data: datetime.strptime(data, "%d-%m-%Y"), reverse=True)
            # recreate tables in sorted order now by reverse engineering the reconvert(convert)
            list_tables_sql_sorted = list()
            for table in list_tables_final_dates:
                sorted_table = self.checker.convert_date(table)
                list_tables_sql_sorted.append(sorted_table)
            i = 0
            for table_name in list_tables_sql_sorted:
                ''''RETRIEVE DATA FROM TABLES'''
                sql = "SELECT * FROM " + table_name
                df = pd.read_sql(sql, connection)
                df.to_excel(writer, sheet_name=list_tables_final_dates[i], index=False)
                i += 1
            '''CUSTOMIZE SHEETS'''
            writer.close()
            self.customize_sheets()
        except:
            messagebox.showerror("INCHIDETI FISIER EXCEL",
                                 "Fisierul Registru_Programari este deschis! Va rog inchideti-l")
            raise Exception("EXCEL file is already opened")
        message_excel = "Baza de date cu pacienti este transferata si disponibila pe {}".format(self.excel_location)
        message_warning = "VA ROG NU EFECTUATI NICI O OPERATIE CAT TIMP VIZUALIZATI FISIERUL EXCEL.\n APLICATIA VA RULA NORMAL DUPA INCHIDEREA EXCELULUI"
        messagebox.showinfo("EXCEL CREATED", message=message_excel)
        messagebox.showinfo("FARA OPERATII", message=message_warning)
        os.system(self.excel_location)

    def customize_sheets(self):
        wb = openpyxl.load_workbook(filename=self.excel_location)
        sheets = wb.sheetnames
        for sheet_name in sheets:
            work_sheet = wb[sheet_name]
            color = openpyxl.styles.colors.Color(rgb="00F5FFDE")
            custom_fill = PatternFill(fill_type="solid", fgColor=color)
            list_index_columns = ["A", "B", "C", "D", "E", "F"]
            # INSERT AN EXTRA COLUMN FOR ID
            work_sheet.insert_cols(idx=1)
            # MAKE THE FIRST COLUMN AS AN ID
            work_sheet["A1"].value = "ID"
            work_sheet["A1"].alignment = Alignment(horizontal="center")
            work_sheet["A1"].font = Font(bold=True)
            for column in list_index_columns:
                # make a custom style
                work_sheet["{}1".format(column)].fill = custom_fill
            # insert the ids from 1 to 16
            for i in range(1, work_sheet.max_row):
                work_sheet["A{}".format(i + 1)].value = i
                work_sheet["A{}".format(i+1)].alignment = Alignment(horizontal="center")
                work_sheet["A{}".format(i+1)].font = Font(bold=True, color="FF0000")
            '''wrap text'''
            # first let's see what is the max value of each row
            list_values_column = []
            for index_column in range(1, len(list_index_columns)):
                list_length_column = list()
                for k in range(1, work_sheet.max_row + 1):
                    # in case it is not completed the cell
                    if work_sheet["{}{}".format(list_index_columns[index_column], k)].value is None:
                        continue
                    else:
                        column_row_length = len(
                            work_sheet["{}{}".format(list_index_columns[index_column], k)].value)
                        list_length_column.append(column_row_length)
                max_value_column = self.get_max_value(list_length_column)
                list_values_column.append(max_value_column)
            # now traverse every column and make than column the width of the max value
            index = 0
            for index_column in range(1, len(list_index_columns)):
                work_sheet.column_dimensions[list_index_columns[index_column]].width = list_values_column[
                                                                                           index] + 10  # make them a bit bigger
                index += 1
            # apply filters => stack overflow code
            full_range = "A1:" + get_column_letter(work_sheet.max_column) \
                         + str(work_sheet.max_row)
            work_sheet.auto_filter.ref = full_range
            wb.save(self.excel_location)
